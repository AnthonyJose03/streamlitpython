import streamlit as st
import pandas as pd
import sqlalchemy
import time
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from io import BytesIO

# configuração do site
st.set_page_config(page_title="Dashboard", layout='wide')

with open("styles.css") as css:
    st.markdown(f'<style>{css.read()}</style>', unsafe_allow_html=True)

# Configuração da conexão com o banco
DB_CONFIG = {
    "usuario": "postgres",
    "senha": "rzEe42WpuRgpwA3O",
    "host": "monstrously-dexterous-arachnid.data-1.use1.tembo.io",
    "porta": "5432",
    "nome_do_banco": "app"
}
engine = sqlalchemy.create_engine(
    f"postgresql+psycopg2://{DB_CONFIG['usuario']}:{DB_CONFIG['senha']}@{DB_CONFIG['host']}:{DB_CONFIG['porta']}/{DB_CONFIG['nome_do_banco']}",
    connect_args={"client_encoding": "utf8"}
)


query_params = st.query_params

id_escola = query_params.get("id_escola", [None])[0]  # Pega o primeiro valor da lista, se houver
id_escola = int(id_escola)

# query de status
status = """
select
	emp.status,
	count(emp.status) as quantidade
from
	emprestimo emp
left outer join
	aluno aln on emp.id_aluno = aln.id
where
    aln.id_escola = %s
group by
	emp.status
"""

qtd_emp_aluno_por_mes = """
select
    aln.nome,
   	l.titulo as livro,
    count(e.id) as quantidade_emprestimos,
    extract(year from e.data_emprestimo) as ano,
    extract(month from e.data_emprestimo) as mes
from
    emprestimo e
left outer join
    aluno aln on e.id_aluno = aln.id
left outer join
	livro l on e.id_livro = l.id
where
    aln.id_escola = %s
group by
    aln.nome, ano, mes, l.titulo
order by
    ano, mes, quantidade_emprestimos desc;
"""

emp_aluno2 = """
select
    aln.nome,
    count(e.id) as quantidade_emprestimos,
    extract(year from e.data_emprestimo) as ano,
    extract(month from e.data_emprestimo) as mes
from
    emprestimo e
left outer join
    aluno aln on e.id_aluno = aln.id
left outer join
	livro l on e.id_livro = l.id
where
    aln.id_escola = %s
group by
    aln.nome, ano, mes
order by
    ano, mes, quantidade_emprestimos desc;
"""

datas_emp = """
select 
	data_emprestimo, 
	data_prevista_entrega , 
	data_entrega 
from 
	emprestimo e
where
    id_escola = %s
"""

livros_m_emprestados = """
select
	l.titulo as titulo,
	count(e.id_livro) as quantidade
from
	emprestimo e
left outer join
	livro l on e.id_livro = l.id
where
    e.id_escola = %s
group by
	l.titulo, e.id_livro
"""

qtd_emprestimo_mes = """
SELECT 
    EXTRACT(YEAR FROM data_emprestimo) AS ano,
    EXTRACT(MONTH FROM data_emprestimo) AS mes,
    COUNT(id) AS quantidade_emprestimos
FROM 
    emprestimo
where
    id_escola = %s
GROUP BY 
    ano, mes
ORDER BY 
    ano DESC, mes DESC;
"""

genero_m_emprestados = """
SELECT
    EXTRACT(YEAR FROM e.data_emprestimo) AS ano,
    EXTRACT(MONTH FROM e.data_emprestimo) AS mes,
    g.nome AS genero,
    COUNT(g.id) AS quantidade
FROM
    rel_livro_genero rlg
LEFT OUTER JOIN
    genero g ON rlg.id_genero = g.id
LEFT OUTER JOIN
    emprestimo e ON rlg.id_livro = e.id_livro
where
    e.id_escola = %s
GROUP BY
    ano, mes, g.nome
ORDER BY
    ano DESC, mes DESC, genero;
"""

qtd_livros = """
select 
    sum(quantidade_livros) as quantidade 
from 
    livro l
where
    id_escola = %s
"""

qtd_livro_emprestado = """
select 
    sum(quantidade_emprestado) as quantidade 
from 
    livro l
where
    id_escola = %s
"""

livros_mais_emprestados_query = """
select
    l.titulo,
    count(e.id) as quantidade_emprestimos,
    extract(year from e.data_emprestimo) as ano,
    extract(month from e.data_emprestimo) as mes
from
    emprestimo e
left outer join
    livro l on e.id_livro = l.id
where
    e.id_escola = %s
group by
    l.titulo, ano, mes
order by
    ano, mes, quantidade_emprestimos desc
"""

livros_disponiveis = """
SELECT 
    l.titulo,
    l.isbn,
    l.quantidade_livros - l.quantidade_emprestado AS quantidade_disponivel
FROM 
    livro l
WHERE 
    l.quantidade_livros > l.quantidade_emprestado
    and
    l.id_escola = %s;
"""

livros_geral = """
SELECT 
    SUM(l.quantidade_emprestado) AS total_livros_emprestados,
    sum(l.quantidade_livros - l.quantidade_emprestado) as livros_disponiveis,
    sum(l.quantidade_livros) as total_geral
FROM 
    livro l
where
    l.id_escola = %s;
"""

quantidade_ano = """
SELECT 
    EXTRACT(YEAR FROM data_emprestimo) AS ano,
    COUNT(id) AS quantidade_emprestimos
FROM 
    emprestimo e
where
    id_escola = %s
GROUP BY 
    ano
ORDER BY 
    ano DESC;
"""

status_alunos = """
select 
	a.nome,
	l.titulo,
    e.data_prevista_entrega as data_prevista,
	e.status 
from 
	emprestimo e
left outer join
	aluno a on e.id = a.id 
left outer join
	livro l on e.id_livro = l.id 
where 
    e.id_escola = %s
order by
    e.status;
"""

acervo_geral = """
select
	titulo,
	isbn,
	data_publicacao,
	autor,
	quantidade_livros,
	quantidade_emprestado,
	localizacao_pratileira
from
	livro
"""

genero_total = """
SELECT
    EXTRACT(YEAR FROM e.data_emprestimo) AS ano,
    g.nome AS genero,
    COUNT(g.id) AS quantidade
FROM
    rel_livro_genero rlg
LEFT OUTER JOIN
    genero g ON rlg.id_genero = g.id
LEFT OUTER JOIN
    emprestimo e ON rlg.id_livro = e.id_livro
WHERE
    e.id_escola = %s
GROUP BY
    ano, g.nome
ORDER BY
    ano DESC, quantidade DESC, genero;
"""

livros_total = """
select
    l.titulo,
    count(e.id) as quantidade_emprestimos,
    extract(year from e.data_emprestimo) as ano
from
    emprestimo e
left outer join
    livro l on e.id_livro = l.id
where
    e.id_escola = %s
group by
    l.titulo, ano
order by
    ano, quantidade_emprestimos desc;
"""

total_alunos = """
select
    aln.nome,
    count(e.id) as quantidade_emprestimos,
    extract(year from e.data_emprestimo) as ano
from
    emprestimo e
left outer join
    aluno aln on e.id_aluno = aln.id
left outer join
	livro l on e.id_livro = l.id
where
    aln.id_escola = %s
group by
    aln.nome, ano
order by
    ano, quantidade_emprestimos desc;
"""


@st.cache_data
def execute_query(query, id_escola):
    return pd.read_sql_query(query, con=engine, params=(id_escola,))

# Transformação de query com o id_escola sendo passado
status = execute_query(status, id_escola)
emp_aluno = execute_query(qtd_emp_aluno_por_mes, id_escola)
datas_emp = execute_query(datas_emp, id_escola)
genero_m_emprestados = execute_query(genero_m_emprestados, id_escola)
qtd_livros = execute_query(qtd_livros, id_escola)
qtd_livro_emprestado = execute_query(qtd_livro_emprestado, id_escola)
livros_m_emprestados = execute_query(livros_m_emprestados, id_escola)
emp_aluno2 = execute_query(emp_aluno2, id_escola)
livros_mais_emprestados = execute_query(livros_mais_emprestados_query, id_escola)
qtd_emprestimo_mes = execute_query(qtd_emprestimo_mes, id_escola)
livros_disponiveis = execute_query(livros_disponiveis, id_escola)
livros_geral = execute_query(livros_geral, id_escola)
quantidade_ano = execute_query(quantidade_ano, id_escola)
status_alunos = execute_query(status_alunos, id_escola)
acervo_geral = execute_query(acervo_geral, id_escola)
genero_total = execute_query(genero_total, id_escola)
livros_total = execute_query(livros_total, id_escola)
total_alunos = execute_query(total_alunos, id_escola)


qtd_emprestimo_mes['ano'] = qtd_emprestimo_mes['ano'].astype(int)
qtd_emprestimo_mes['mes'] = qtd_emprestimo_mes['mes'].astype(int)
emp_aluno['ano'] = emp_aluno['ano'].astype(int)
qtd_emprestimo_mes['data'] = pd.to_datetime(qtd_emprestimo_mes['ano'].astype(str) + '-' + qtd_emprestimo_mes['mes'].astype(str) + '-01', format='%Y-%m-%d')

df = qtd_emprestimo_mes.sort_values(by='data', ascending=True)
# Mapeamento de meses
meses = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
    7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}
meses_2 = {
    1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
    7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
}

# Mapeando o número do mês para o nome do mês
emp_aluno['mes'] = emp_aluno['mes'].map(meses)
emp_aluno2['mes'] = emp_aluno2['mes'].map(meses)
livros_mais_emprestados['mes'] = livros_mais_emprestados['mes'].map(meses)
qtd_emprestimo_mes['mes'] = qtd_emprestimo_mes['mes'].map(meses)
df['mes'] = df['mes'].map(meses_2)
genero_m_emprestados['mes'] = genero_m_emprestados['mes'].map(meses)


# Função para converter o DataFrame para Excel
def convert_df_to_excel(df):
    # Criar um arquivo Excel em memória
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    return output.getvalue()

# dicionario do status   
dicionario = {
    'EM_DIA': 'Em dia',
    'EM_ATRASO': 'Em atraso'
}

# renomeando os status do empréstimo
status = status.replace(dicionario)

# Titulo da biblioteca
locacao1, locacao2, locacao3 = st.columns(3)
with locacao2:
    st.markdown("""
    <style>
        .title-container {
            text-align: center;
            margin-top: 10px;
        }
        .title-container h1 {
            font-size: 48px;
            font-weight: bold;
        }
        .title-container p {
            font-size: 25px;
            color: grey;
            font-style: italic;
        }
    </style>
    <div class="title-container">
        <h1>Conexão Literária</h1>
        <p>A conexão da literatura com o aluno.</p>
    </div>
    """, unsafe_allow_html=True)
    
    
# Definir as opções de filtros (ano e mês)
ano_opcoes = emp_aluno['ano'].unique()
mes_opcoes = emp_aluno['mes'].unique()

st.divider()

mes = mes_opcoes.tolist()



col1, col2, col3, col4, col5, col6 = st.columns(6)
with col3:
    ano_selecionado = st.selectbox('Selecione o ano', ano_opcoes)
# Adicionando a opção "Todos os meses"
with col4:
    mes_selecionado = st.selectbox(
    "Selecione o mês", ["Todos os meses"] + mes
    )

# Livro mais emprestado nos meses
# - agrupando por ano, mês e livro
emp_por_mes = emp_aluno.groupby(['ano', 'mes', 'livro']).size().reset_index(name='total_emprestimo')
# - ordenando pelo total de empréstimos
emp_por_mes = emp_por_mes.sort_values(by='total_emprestimo', ascending=False)

# Filtrando por ano e mês selecionado
if mes_selecionado != "Todos os meses":
    livros_mais_emprestados_mes = livros_mais_emprestados.loc[
        (livros_mais_emprestados['ano'] == ano_selecionado) & (livros_mais_emprestados['mes'] == mes_selecionado)
    ]
else:
    livros_mais_emprestados_mes = livros_mais_emprestados.loc[
        livros_mais_emprestados['ano'] == ano_selecionado
    ]

# Verificar os livros com o maior número de empréstimos
max_emprestimos = livros_mais_emprestados_mes["quantidade_emprestimos"].max()
# print(max_emprestimos)

# Filtrar os livros com a quantidade máxima de empréstimos
livros_mais_emprestados_max = livros_mais_emprestados_mes[
    livros_mais_emprestados_mes["quantidade_emprestimos"] == max_emprestimos
]

# Verificar os livros com o maior número de empréstimos durante o ano
max_emprestimos_ano = livros_mais_emprestados_mes.groupby('titulo')['quantidade_emprestimos'].sum().max()

# Filtrar os livros com a quantidade máxima de empréstimos no ano
livros_mais_emprestados_max_ano = livros_mais_emprestados_mes.groupby('titulo')['quantidade_emprestimos'].sum().reset_index()

# Filtra os livros que têm o maior número de empréstimos no ano
livros_mais_emprestados_max_ano = livros_mais_emprestados_max_ano[
    livros_mais_emprestados_max_ano["quantidade_emprestimos"] == max_emprestimos_ano
]

print(livros_mais_emprestados_max)

# se selecionar a opção de todos os meses
if mes_selecionado == 'Todos os meses':
    if len(livros_mais_emprestados_max_ano) > 1:
        barr1, barr2, barr3 = st.columns([2, 6, 2])  # Usando proporção maior para a coluna do meio
        with barr2:
            # Estilização do título
            st.markdown(f"""
                <div class="title-container" style="text-align: center; margin-bottom: 20px;">
                    <h2 style="color: black;
                        font-family: 'Arial', sans-serif;">Livros mais emprestados no ano:</h2>
                </div>
            """, unsafe_allow_html=True)

            coluns1, coluns2 = st.columns(2)
            with coluns1:
                # Exibe o primeiro card do ano
                livro_1 = livros_mais_emprestados_max_ano.iloc[0]
                st.markdown(f"""
                    <div style="background-color: #f9f9f9; 
                        padding: 30px; 
                        border-radius: 15px; 
                        text-align: center; 
                        font-size: 24px; 
                        border: 3px solid #ddd; 
                        box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.1); 
                        width: 100%; 
                        max-width: 400px; 
                        margin: auto;">
                        <b style="font-size: 30px; 
                            color: #333;">{livro_1["titulo"]}</b><br><br>
                        <i style="font-size: 26px; 
                            color: #777;">Total de Empréstimos: {livro_1["quantidade_emprestimos"]}</i>
                    </div>
                """, unsafe_allow_html=True)
            with coluns2:
                # Exibe o segundo card do ano
                livro_2 = livros_mais_emprestados_max_ano.iloc[1]
                st.markdown(f"""
                    <div style="background-color: #f9f9f9; 
                        padding: 30px; 
                        border-radius: 15px; 
                        text-align: center; 
                        font-size: 24px; 
                        border: 3px solid #ddd; 
                        box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.1); 
                        width: 100%; 
                        max-width: 400px; 
                        margin: auto;">
                        <b style="font-size: 30px; 
                            color: #333;">{livro_2["titulo"]}</b><br><br>
                        <i style="font-size: 26px; 
                            color: #777;">Total de Empréstimos: {livro_2["quantidade_emprestimos"]}</i>
                    </div>
                """, unsafe_allow_html=True)

    else:
        # Se houver apenas um livro mais emprestado no ano, exibe o card normal
        barr1, barr2, barr3 = st.columns([2, 6, 2])  # Usando proporção maior para a coluna do meio
        with barr2:
            st.markdown(f"""
                <div class="title-container" style="text-align: center; margin-bottom: 20px;">
                    <h2 style="color: black;
                        font-family: 'Arial', sans-serif;">Livro mais emprestado no ano:</h2>
                </div>
            """, unsafe_allow_html=True)

            # Card estilizado
            st.markdown(f"""
                <div style="background-color: #f9f9f9; 
                    padding: 30px; 
                    border-radius: 15px; 
                    text-align: center; 
                    font-size: 24px; 
                    border: 3px solid #ddd; 
                    box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.1); 
                    width: 100%; 
                    max-width: 400px; 
                    margin: auto;">
                    <b style="font-size: 30px; 
                        color: #333;">{livros_mais_emprestados_max_ano.iloc[0]["titulo"]}</b><br><br>
                    <i style="font-size: 26px; 
                        color: #777;">Total de Empréstimos: {livros_mais_emprestados_max_ano.iloc[0]["quantidade_emprestimos"]}</i>
                </div>
            """, unsafe_allow_html=True)
else:
    if len(livros_mais_emprestados_max) > 1:
        barr1, barr2, barr3 = st.columns([2, 6, 2])  # Usando proporção maior para a coluna do meio
        with barr2:
            # Estilização do título
            st.markdown(f"""
                <div class="title-container" style="text-align: center; margin-bottom: 20px;">
                    <h2 style="color: black;
                        font-family: 'Arial', sans-serif;">Livros mais emprestados no mês de {mes_selecionado}:</h2>
                </div>
            """, unsafe_allow_html=True)

            coluns1, coluns2 = st.columns(2)
            with coluns1:
                # Exibe o primeiro card
                livro_1 = livros_mais_emprestados_max.iloc[0]
                st.markdown(f"""
                    <div style="background-color: #f9f9f9; 
                        padding: 30px; 
                        border-radius: 15px; 
                        text-align: center; 
                        font-size: 24px; 
                        border: 3px solid #ddd; 
                        box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.1); 
                        width: 100%; 
                        max-width: 400px; 
                        margin: auto;">
                        <b style="font-size: 30px; 
                            color: #333;">{livro_1["titulo"]}</b><br><br>
                        <i style="font-size: 26px; 
                            color: #777;">Total de Empréstimos: {livro_1["quantidade_emprestimos"]}</i>
                    </div>
                """, unsafe_allow_html=True)
            with coluns2:
                # Exibe o segundo card
                livro_2 = livros_mais_emprestados_max.iloc[1]
                st.markdown(f"""
                    <div style="background-color: #f9f9f9; 
                        padding: 30px; 
                        border-radius: 15px; 
                        text-align: center; 
                        font-size: 24px; 
                        border: 3px solid #ddd; 
                        box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.1); 
                        width: 100%; 
                        max-width: 400px; 
                        margin: auto;">
                        <b style="font-size: 30px; 
                            color: #333;">{livro_2["titulo"]}</b><br><br>
                        <i style="font-size: 26px; 
                            color: #777;">Total de Empréstimos: {livro_2["quantidade_emprestimos"]}</i>
                    </div>
                """, unsafe_allow_html=True)

    else:
        # - puxando o o primeiro index dos emprestimos ordenados
        livro_mais_emprestado = livros_mais_emprestados.loc[livros_mais_emprestados['mes'] == mes_selecionado].iloc[0]

        # Se houver apenas um livro mais emprestado, exibe o card normal
        barr1, barr2, barr3 = st.columns([2, 6, 2])  # Usando proporção maior para a coluna do meio
        with barr2:
            st.markdown(f"""
                <div class="title-container" style="text-align: center; margin-bottom: 20px;">
                    <h2 style="color: black;
                        font-family: 'Arial', sans-serif;">Livro mais emprestado no mês de {mes_selecionado}:</h2>
                </div>
            """, unsafe_allow_html=True)
            
            
            # Card estilizado
            st.markdown(f"""
                <div style="background-color: #f9f9f9; 
                    padding: 30px; 
                    border-radius: 15px; 
                    text-align: center; 
                    font-size: 24px; 
                    border: 3px solid #ddd; 
                    box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.1); 
                    width: 100%; 
                    max-width: 400px; 
                    margin: auto;">
                    <b style="font-size: 30px; 
                        color: #333;">{livro_mais_emprestado["titulo"]}</b><br><br>
                    <i style="font-size: 26px; 
                        color: #777;">Total de Empréstimos: {livro_mais_emprestado["quantidade_emprestimos"]}</i>
                </div>
            """, unsafe_allow_html=True)


st.divider()

# Extraindo os valores da consulta
total_livros_emprestados = livros_geral['total_livros_emprestados'].iloc[0]
livros_disponiveis = livros_geral['livros_disponiveis'].iloc[0]
total_geral = livros_geral['total_geral'].iloc[0]

# criando o dataframe para o grafico de pizza
distribuicao_livros = pd.DataFrame({
    'status': ['Livros Disponíveis', 'Livros Emprestados'],
    'quantidade': [livros_disponiveis, total_livros_emprestados]
})    

# renomeando as colunas
status = status.rename(columns={
    'status': 'Status',
    'quantidade': 'Quantidade'
})
distribuicao_livros = distribuicao_livros.rename(columns={
    'status': 'Status',
    'quantidade': 'Quantidade'
})
emp_aluno = emp_aluno.rename(columns={
    'nome': 'Nome'
})
genero_m_emprestados = genero_m_emprestados.rename(columns={
    'genero': 'Gênero',
    'quantidade': 'Quantidade'
})
livros_m_emprestados = livros_m_emprestados.rename(columns={
    'titulo': 'Título',
    'quantidade': 'Quantidade'
})
qtd_emprestimo_mes = qtd_emprestimo_mes.rename(columns={
    'mes': 'Mês',
    'quantidade_emprestimos': 'Quantidade',
    'ano': 'Ano'
})
df = df.rename(columns={
    'mes': 'Mês',
    'quantidade_emprestimos': 'Quantidade',
    'ano': 'Ano',
    'data': 'Data'
})
genero_total = genero_total.rename(columns={
    'quantidade': 'Quantidade',
    'genero': 'Gênero'
})


# grafico de pizza (status do emprestimo)
fig1 = px.pie(status, names='Status',
             values='Quantidade',
             title='Status do Emprestimo',
             color='Status',
             color_discrete_map={'Em dia':'#65558F',
                                 'Em atraso':'#483D8B'})
fig1.update_traces(textfont_size=20,
                   hovertemplate = '<b>%{label}</b><br>Quantidade: %{value}<extra></extra>',
                   hoverlabel = dict(
                       font_size = 14
                   ))

# parâmetros do gráfico
fig1.update_layout(
    title=dict(
        font=dict(
            size=24,
            color="black"    # Tamanho da fonte do título
            )        
        ),
    legend=dict(
        font=dict(
            size=15,
            color="black"# Tamanho da fonte da legenda (em pixels)
        )
    )
)

# gráfico de pizza (distribuição dos livros)
fig2 = px.pie(distribuicao_livros, names='Status',
             values='Quantidade',
             title='Distribuição dos Livros',
             color='Status', 
             color_discrete_map={'Livros Disponíveis': '#4E358E',
                                 'Livros Emprestados': '#751AAA'})

# parâmetros do gráfico
fig2.update_layout(
    title=dict(
        font=dict(
            size=24,
            color="black"    # Tamanho da fonte do título
            )        
        ),
    legend=dict(
        font=dict(
            size=15,
            color="black"# Tamanho da fonte da legenda (em pixels)
        )
    )
)

# Sobrepondo o valor das fatias
fig2.update_traces(textposition='auto',
                    textfont=dict(size=20,
                                  color="white"),
                    hovertemplate = '<b>%{label}</b><br>Quantidade: %{value}<extra></extra>',
                    hoverlabel = dict(
                       font_size = 14
                   ))

# Calcular o total de cada status
em_dia = status[status['Status'] == 'Em dia']['Quantidade'].sum()
pendente = status[status['Status'] == 'Pendente']['Quantidade'].sum()
em_atraso = status[status['Status'] == 'Em atraso']['Quantidade'].sum()
em_dia_real = em_dia+pendente

# Calcular o acervo de livros disponiveis e emprestados
qtd_livros = qtd_livros.reset_index()
qtd_livro_emprestado = qtd_livro_emprestado.reset_index()
qtd_livros = qtd_livros['quantidade']
qtd_livro_emprestado = qtd_livro_emprestado['quantidade']
disponiveis = qtd_livros-qtd_livro_emprestado

# colunas do grafico de pizza
col1, col2, col3, col4 = st.columns([5, 4, 5.7, 4])
with col1:
    with st.container(border=True):
        # plotando o gráfico de status
        st.plotly_chart(fig1)
with col2:
    # totalizadores do gráfico de status
    container = st.container(border=True)
    container.metric(label="**Total em dia**", value=em_dia_real)
    container.metric(label="**Total em atraso**", value=em_atraso)
    
    # Função para ajustar o arquivo Excel gerado
    def ajustar_excel(df):
        # Salvar o DataFrame em um arquivo Excel em memória (sem salvar no disco)
        output = BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)

        # Abrir o arquivo Excel gerado
        wb = load_workbook(output)
        ws = wb.active
        
        # Criar um estilo de data
        data_style = NamedStyle(name="data_style", number_format="YYYY-MM-DD")
        
        # Ajustar largura da coluna de datas (supondo que a data esteja na coluna A)
        ws.column_dimensions['A'].width = 20  # Ajustar a largura da coluna A
        ws.column_dimensions['B'].width = 20  # Ajustar a largura da coluna B
        ws.column_dimensions['C'].width = 20  # Ajustar a largura da coluna C
                
        # Aplicar o estilo de data para todas as células da coluna A
        for cell in ws['A']:
            cell.style = data_style
        
        for cell in ws['B']:
            cell.style = data_style
        
        for cell in ws['C']:
            cell.style = data_style
        
        # Salvar o arquivo Excel com as modificações
        output.seek(0)
        wb.save(output)
        output.seek(0)
        
        return output

    # Ajustar o arquivo Excel
    status_excel = ajustar_excel(status_alunos)

    # Download do arquivo Excel ajustado
    st.download_button(
        label="Baixar Tabela de Status",
        data=status_excel,
        file_name="tabela_status.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
with col3:
    with st.container(border=True):
        # plotando o grafico de distribuição
        st.plotly_chart(fig2)
with col4:
    # totalizadores do gráfico de distribuição
    container2 = st.container(border=True)
    container2.metric(label="**Acervo Geral**", value=qtd_livros)
    container2.metric(label="**Livros Disponiveis**", value=disponiveis)
    container2.metric(label="**Livros Emprestados**", value=qtd_livro_emprestado)
    
    excel = convert_df_to_excel(acervo_geral)
    st.download_button(
        label="Baixar Acervo Geral",
        data=excel,
        file_name='acervo_geral.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )




# Gráfico de barra (emprestimos por mes)
fig3 = px.bar(df,
              x='Mês',
              y='Quantidade', 
              title='Empréstimos anual/mensal',
              color_discrete_sequence=["#65558F"],
              text='Quantidade')

fig3.update_layout(
    title=dict(
        font=dict(
            size=24,
            color="black"      # Tamanho da fonte do título
        )
    ),
    xaxis=dict(
        title=dict(
            font=dict(
                size=18,
                color="black"  # Tamanho da fonte do eixo X
            )
        ),
        tickfont=dict(
            size=16,
            color="black"      # Tamanho da fonte dos rótulos do eixo X
        )
    ),
    yaxis=dict(
        title=dict(
            font=dict(
                size=18,
                color="black"  # Tamanho da fonte do título do eixo Y
            )
        ),
        tickfont=dict(
            size=16,
            color="black"      # Tamanho da fonte dos rótulos do eixo Y
        )
    )
)    

fig3.update_layout(
    xaxis=dict(
        dtick=2  # Intervalo entre as marcas no eixo x
    ),
    yaxis=dict(
        dtick=2  # Intervalo entre as marcas no eixo y
    )
)

fig3.update_layout(
    width=800,  # Largura
    height=500  # Altura
)

fig3.update_layout(
    xaxis=dict(
        tickmode='array',  # Garantir que todos os ticks sejam mostrados
        tickvals=df['Mês'],  # Certifica que os meses são usados
    )
)

max_y1 = df['Quantidade'].max()


fig3.update_layout(
    yaxis=dict(
        range=[0, max_y1 * 1.1]  # Ajuste do limite superior (aumentando 10% do valor máximo)
    )
)

config={
    'displayModeBar': 'hover',  # A barra de ferramentas aparece ao passar o mouse
    'modeBarButtonsToRemove': [
        'zoom', 'pan', 'resetScale', 'hover', 'zoomIn', 'zoomOut', 'resetAxes', 
        'lasso2d', 'select2d', 'zoom2d', 'pan2d', 'orbitRotation', 'tableRotation', 
        'save', 'zoom2d', 'sendDataToCloud', 'toggleSpikelines', 'toggleHover', 
        'resetView', 'toggleFullscreen'  # Remove todos os botões, incluindo o de tela cheia
    ],
    'displaylogo': False,  # Remove o logo do Plotly da barra de ferramentas
    'modeBarButtonsToAdd': ['downloadImage']  # Adiciona apenas o botão de download
}

# sobrepondo as barras com os valores
fig3.update_traces(textposition='outside',
                    textfont=dict(size=18,
                                  color='black'),
                    hovertemplate = '<b>%{label}</b><br>Quantidade: %{value}<extra></extra>',
                    hoverlabel = dict(
                       font_size = 14
                   ))

# totalizador dos emprestimos anuais
total_emprestimos = quantidade_ano['quantidade_emprestimos'].sum()
fig3.add_annotation(
    x=1,  # posição x (canto direito)
    y=1.25,  # posição y (canto superior)
    text=f"{total_emprestimos} empréstimos",
    showarrow=False,
    font=dict(size=20, color="black"),
    align="left",
    xref="paper",  # Usando coordenadas relativas ao espaço da figura
    yref="paper",  # Usando coordenadas relativas ao espaço da figura
    bgcolor="white"  # Fundo branco para destacar
)

# plotando o gráfico de barras (emprestimos por mês)
with st.container(border=True):
    st.plotly_chart(fig3, use_container_width=True, config=config)


# Filtrando os dados com base nas seleções de ano e mês
generos_filtrados = genero_m_emprestados[
    (genero_m_emprestados['ano'] == ano_selecionado) & 
    (genero_m_emprestados['mes'] == mes_selecionado)
]

custom_colors = ['#BEA6FF', '#AF54E4', '#A18AC2', '#8E74D2',
                 '#751AAA', '#DCCEFF', '#A887FF', '#4E358E',
                 '#7B7097', '#824BA7']


if mes_selecionado == 'Todos os meses':
    fig4 = px.bar(genero_total, x='Gênero',
            y='Quantidade',
            title='Emprestimos por gênero/anual',
            color_discrete_sequence=custom_colors,
            color='Gênero',
            text='Quantidade')
    fig4.update_layout(
        title=dict(
            font=dict(
                size=24             # Tamanho da fonte do título
            )
        ),
        xaxis=dict(
            title=dict(
                font=dict(
                    size=18,
                    color="black"   # Tamanho da fonte do título do eixo X
                )
            ),
            tickfont=dict(
                size=16,
                color="black"
                                    # Tamanho da fonte dos rótulos do eixo X
            )
        ),
        yaxis=dict(
            title=dict(
                font=dict(
                    size=18,
                    color="black"   # Tamanho da fonte do título do eixo Y
                )
            ),
            tickfont=dict(
                size=16,
                color="black"       # Tamanho da fonte dos rótulos do eixo Y
            )
        ),
        legend=dict(
            font=dict(
                size=14,
                color="black"       # Tamanho da fonte da legenda
            )
        )
    )

    # sobrepondo as barras do gráfico de gêneros
    fig4.update_traces(textposition='outside',
                        textfont=dict(size=18,
                                    color='black'),
                        hovertemplate = '<b>%{label}</b><br>Quantidade: %{value}<extra></extra>',
                        hoverlabel = dict(
                            font_size = 14
                    ))

    max_y2 = genero_total['Quantidade'].max()

    fig4.update_layout(
        xaxis=dict(
            dtick=2  # Intervalo entre as marcas no eixo x
        ),
        yaxis=dict(
            dtick=2  # Intervalo entre as marcas no eixo y
        )
    )

    fig4.update_layout(
        xaxis=dict(
            tickmode='array',  # Garantir que todos os ticks sejam mostrados
            tickvals=genero_total['Gênero'],  # Certifica que os meses são usados
        )
    )

    fig4.update_layout(
        width=800,  # Largura
        height=500  # Altura
    )

    fig4.update_layout(
        yaxis=dict(
            range=[0, max_y2 * 1.1]  # Ajuste do limite superior (aumentando 10% do valor máximo)
        )
    )
else:
    fig4 = px.bar(generos_filtrados, x='Gênero',
                y='Quantidade',
                title='Emprestimos por gênero/mensal',
                color_discrete_sequence=custom_colors,
                color='Gênero',
                text='Quantidade')
    fig4.update_layout(
        title=dict(
            font=dict(
                size=24             # Tamanho da fonte do título
            )
        ),
        xaxis=dict(
            title=dict(
                font=dict(
                    size=18,
                    color="black"   # Tamanho da fonte do título do eixo X
                )
            ),
            tickfont=dict(
                size=16,
                color="black"
                                    # Tamanho da fonte dos rótulos do eixo X
            )
        ),
        yaxis=dict(
            title=dict(
                font=dict(
                    size=18,
                    color="black"   # Tamanho da fonte do título do eixo Y
                )
            ),
            tickfont=dict(
                size=16,
                color="black"       # Tamanho da fonte dos rótulos do eixo Y
            )
        ),
        legend=dict(
            font=dict(
                size=14,
                color="black"       # Tamanho da fonte da legenda
            )
        )
    )

    # sobrepondo as barras do gráfico de gêneros
    fig4.update_traces(textposition='outside',
                        textfont=dict(size=18,
                                    color='black'))

    max_y2 = generos_filtrados['Quantidade'].max()

    fig4.update_layout(
        xaxis=dict(
            dtick=2  # Intervalo entre as marcas no eixo x
        ),
        yaxis=dict(
            dtick=2  # Intervalo entre as marcas no eixo y
        )
    )

    fig4.update_layout(
        xaxis=dict(
            tickmode='array',  # Garantir que todos os ticks sejam mostrados
            tickvals=generos_filtrados['Gênero'],  # Certifica que os gêneros são usados
        )
    )

    fig4.update_layout(
        width=800,  # Largura
        height=500  # Altura
    )

    fig4.update_layout(
        yaxis=dict(
            range=[0, max_y2 * 1.1]  # Ajuste do limite superior (aumentando 10% do valor máximo)
        )
    )

# plotando o grafico de barras (gêneros mais emprestados)
with st.container(border=True):
    st.plotly_chart(fig4, use_container_width=True, config=config)    
    
# Estilização das tabelas
st.markdown(
    """
    <style>
    .center-table {
        display: flex;
        justify-content: left;
        align-items: left;
    }
    table {
        font-size: 20px;
        width: 75%; /* Ajuste o tamanho da tabela aqui */
        border-radius: 10px;
        border-collapse: collapse;
        border: 6px #080808;
        overflow: hidden;
    }
    th, td {
        text-align: left;
        border: 4px solid #ddd;
        padding: 8px;
    }
    th {
        color: white;
        background-color: #65558F;
        font-weight: bold;
    }
    /* Linhas com fundo branco e roxo claro alternado */
    tr:nth-child(odd) {
        background-color: #65558F17; /* Cinza claro */
    }
    tr:nth-child(even) {
        background-color: white; /* Branco */
    }
    </style>
    """,
    unsafe_allow_html=True
    )

# Tabelas de livros e alunos
bar1, bar2 = st.columns(2)
with bar1:
    if mes_selecionado == 'Todos os meses':
        livros_total = livros_total.rename(columns={
            'titulo': 'Título',
            'quantidade_emprestimos': 'n° emprestimos'
        })
        livros_total = livros_total[['Título', 'n° emprestimos']].nlargest(10, 'n° emprestimos')
        livros_tabela = livros_total
        livros_total = livros_total.to_html(index=False)        
        st.write("## Livros mais emprestados")
        st.write("Mês:", mes_selecionado)
        st.markdown(f'<div class="center-table">{livros_total}</div>', unsafe_allow_html=True)
        
        # Download dos livros
        excel = convert_df_to_excel(livros_tabela)
        st.download_button(
        label="Baixar Tabela",
        data=excel,
        file_name='tabela_livros.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    else:
        livros_mais_emprestados_filtrado =  livros_mais_emprestados[(livros_mais_emprestados['ano'] == ano_selecionado) & (livros_mais_emprestados['mes'] == mes_selecionado)]
        # Ranking dos livros mais emprestados
        livros_mais_emprestados_filtrado = livros_mais_emprestados_filtrado.rename(columns={
            'titulo': 'Título',
            'quantidade_emprestimos': 'n° emprestimos'
        })
        livros_mais_emprestados_filtrado = livros_mais_emprestados_filtrado[['Título', 'n° emprestimos']].nlargest(10, 'n° emprestimos')
        tabela_livros = livros_mais_emprestados_filtrado
        livros_mais_emprestados_filtrado = livros_mais_emprestados_filtrado.to_html(index=False)
        st.write("## Livros mais emprestados")
        st.write("Mês:", mes_selecionado)
        st.markdown(f'<div class="center-table">{livros_mais_emprestados_filtrado}</div>', unsafe_allow_html=True)
        
        
        # Download dos livros
        excel = convert_df_to_excel(tabela_livros)
        st.download_button(
        label="Baixar Tabela",
        data=excel,
        file_name='tabela_livros.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

with bar2:
    if mes_selecionado == 'Todos os meses':
        total_alunos = total_alunos.rename(columns={
            'nome': 'Nome',
            'quantidade_emprestimos': 'n° emprestimos'
        })
        total_alunos = total_alunos[['Nome', 'n° emprestimos']].nlargest(10, 'n° emprestimos')
        tabela_total_alunos = total_alunos
        total_alunos = total_alunos.to_html(index=False, escape=False)
        st.write("## Ranking de alunos")
        st.write("Mês:", mes_selecionado)
        st.markdown(f'<div class="center-table">{total_alunos}</div>', unsafe_allow_html=True)
        
        excel = convert_df_to_excel(tabela_total_alunos)
        st.download_button(
            label="Baixar Tabela",
            data=excel,
            file_name='tabela_alunos.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
     )
            
    else:
        # Ranking dos alunos que mais pegaram livros
        emp_aluno_filtrado = emp_aluno2[(emp_aluno2['ano'] == ano_selecionado) & (emp_aluno2['mes'] == mes_selecionado)]
        emp_aluno_filtrado = emp_aluno_filtrado.rename(columns={
            'nome': 'Nome',
            'quantidade_emprestimos': 'n° emprestimos'
        })
        emp_aluno_filtrado = emp_aluno_filtrado[['Nome', 'n° emprestimos']].nlargest(10, 'n° emprestimos')
        tabela_alunos = emp_aluno_filtrado
        emp_aluno_filtrado = emp_aluno_filtrado.to_html(index=False, escape=False)
        st.write("## Ranking de alunos")
        st.write("Mês:", mes_selecionado)
        st.markdown(f'<div class="center-table">{emp_aluno_filtrado}</div>', unsafe_allow_html=True)
    
        excel = convert_df_to_excel(tabela_alunos)
        st.download_button(
            label="Baixar Tabela",
            data=excel,
            file_name='tabela_alunos.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
