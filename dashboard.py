import streamlit as st
import pandas as pd
import plotly.express as px
import matplotlib.pyplot as plt
import base64
from wordcloud import WordCloud
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from io import BytesIO
from sql import *

# configuração do site
st.set_page_config(page_title="Dashboard", layout="wide")

with open("styles.css") as css:
    st.markdown(f"<style>{css.read()}</style>", unsafe_allow_html=True)
    
query_params = st.query_params

id_escola = query_params.get("id_escola", [None])[0]
id_escola = int(id_escola)

@st.cache_data
def execute_query(query, id_escola):
    return pd.read_sql_query(query, con=engine, params=(id_escola,))

def to_base64(file_data):
    return base64.b64encode(file_data).decode("utf-8")

def ajustar_excel(df):
    output = BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # Criar um estilo de data
    data_style = NamedStyle(name="data_style", number_format="DD-MM-YYYY")

    ws.column_dimensions["A"].width = 20 
    ws.column_dimensions["B"].width = 20 
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 20

    for cell in ws["A"]:
        cell.style = data_style

    for cell in ws["B"]:
        cell.style = data_style

    for cell in ws["C"]:
        cell.style = data_style

    output.seek(0)
    wb.save(output)
    output.seek(0)

    return output

def tabela_download(df, file_name):
    excel = ajustar_excel(df)
    excel_base64 = to_base64(excel.read())

    imagem_download = f"""
            <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" 
                download="{file_name}" 
                style="position: absolute; bottom: -20px; right: 1%; text-decoration: none;">
                <img src="data:image/svg+xml;base64,{encoded_image2}" alt="Ícone de download" style="width: 100%; height: 40px;" />
        </a>
            </div>
        </div>
    """

    st.markdown(imagem_download, unsafe_allow_html=True)

def totalizador(label, valor, df, file_name):
    excel = ajustar_excel(df)
    excel_base64 = to_base64(excel.read())

    totalizador_imagem = f"""
        <div style="background: #8E58BF; padding: 2px; border-radius: 15px; width: 92%;  
                    display: flex; justify-content: space-between; align-items: center; text-align: left;
                    padding-left: 19px"> 
            <div style="flex: 3">
                <h3 style="font-size: 22px; color: white; font-weight: lighter;  position: relative; top: 11px">{label}</h3>
                <p style="font-size: 42px; color: white; font-weight: bold; position: relative; left: 5px; top: -8px">
                          {valor}</p>
            </div>
            <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" 
                download="{file_name}" 
                style="position: absolute; bottom: 0px; right: 12%; text-decoration: none;">
                <img src="data:image/svg+xml;base64,{encoded_image}" alt="Ícone de download" style="width: 100%; height: 40px;" />
        </a>
            </div>
        </div>
    """

    st.markdown(totalizador_imagem, unsafe_allow_html=True)

@st.cache_data
def gerar_nuvem_emprestimos(emprestimo_aluno, ano_selecionado, mes_selecionado):

    emp_por_mes = (
        emprestimo_aluno.groupby(["ano", "mes", "livro"])
        .size()
        .reset_index(name="total_emprestimo")
    )
    emp_por_mes = emp_por_mes.sort_values(by="total_emprestimo", ascending=False)

    if not mes_selecionado or "Todos os meses" in mes_selecionado:
        livros_mais_emprestados_mes = emp_por_mes.loc[
            emp_por_mes["ano"] == ano_selecionado
        ]
    else:
        livros_mais_emprestados_mes = emp_por_mes.loc[
            (emp_por_mes["ano"] == ano_selecionado)
            & (emp_por_mes["mes"].isin(mes_selecionado))
        ]

    if livros_mais_emprestados_mes.empty:
        st.warning("Não há dados de empréstimos para os filtros selecionados.")
        return None, None

    livros_mais_emprestados_total = (
        livros_mais_emprestados_mes.groupby("livro")["total_emprestimo"]
        .sum()
        .reset_index()
    )

    livros_freq = dict(
        zip(
            livros_mais_emprestados_total["livro"],
            livros_mais_emprestados_total["total_emprestimo"],
        )
    )

    if not livros_freq:
        st.warning("Não há livros suficientes para gerar a nuvem de palavras.")
        return None, None

    wordcloud = WordCloud(
        width=2000, height=800, max_words=30, background_color="white"
    ).generate_from_frequencies(livros_freq)

    return wordcloud, livros_mais_emprestados_mes

def plot_emprestimos_mensais():
    fig = px.bar(
        emprestimo_ano,
        x="Mês",
        y="Quantidade",
        title="Empréstimos mensais",
        color_discrete_sequence=["#8E58BF"],
        text="Quantidade",
    )

    fig.update_layout(
        title_font=dict(size=24, color="black"),
        xaxis_title_font=dict(size=22, color="black"),
        yaxis_title_font=dict(size=22, color="black"),
        xaxis_tickfont=dict(size=22, color="black"),
        yaxis_tickfont=dict(size=22, color="black"),
        width=800,
        height=500,
        xaxis_tickmode="array",
        xaxis_tickvals=emp_mes["Mês"],
        yaxis_range=[0, emprestimo_ano["Quantidade"].max() * 1.1],
    )

    fig.update_traces(
        textposition="outside",
        textfont=dict(size=20, color="black"),
        hovertemplate="<b>%{label}</b><br>Quantidade: %{value}<extra></extra>",
        hoverlabel=dict(font_size=20),
        marker=dict(line=dict(width=0.1), cornerradius=10),
    )

    total_emprestimos = quantidade_ano[quantidade_ano["ano"] == ano_selecionado][
        "quantidade_emprestimos"
    ].sum()
    fig.add_annotation(
        x=1,
        y=1.25,
        text=f"{total_emprestimos} empréstimos em {ano_selecionado}",
        showarrow=False,
        font=dict(size=24, color="black"),
        align="left",
        xref="paper",
        yref="paper",
        bgcolor="white",
    )

    return fig

def grafico_emprestimo_genero(custom_colors):
    fig = px.bar(
        genero_total_filtrado,
        x="Gênero",
        y="Quantidade",
        title="Empréstimos por gênero",
        color_discrete_sequence=custom_colors,
        color="Gênero",
        text="Quantidade",
        height=800,
    )

    fig.update_layout(
        title=dict(font=dict(size=26)),
        xaxis=dict(
            title=dict(font=dict(size=22, color="black")),
            tickfont=dict(size=22, color="black"),
            tickmode="array",
            tickvals=genero_total_filtrado["Gênero"],
        ),
        yaxis=dict(
            title=dict(font=dict(size=22, color="black")),
            tickfont=dict(size=22, color="black"),
            range=[0, genero_total_filtrado["Quantidade"].max() * 1.1],
        ),
        legend=dict(font=dict(size=20, color="black")),
        width=800,
        height=500,
    )

    fig.update_traces(
        marker=dict(line=dict(width=0.1), cornerradius=10),
        textposition="outside",
        textfont=dict(size=22, color="black"),
        hovertemplate="<b>%{label}</b><br>Quantidade: %{value}<extra></extra>",
        hoverlabel=dict(font_size=18),
    )

    return fig

def grafico_rank_livros(dados):
    fig = px.bar(
        dados,
        x="N° empréstimos",
        y="Título",
        title="Ranking de livros",
        orientation='h',
        color_discrete_sequence=["#8E58BF"],
        text="N° empréstimos",
        height=600,
    )

    fig.update_layout(
        title=dict(font=dict(size=26)),
        xaxis=dict(
            title=dict(font=dict(size=22, color="black")),
            showgrid=True,
            tickfont=dict(size=22, color="black"),
            tickmode="array",
            tickvals=dados["Título"],
        ),
        yaxis=dict(
            tickfont=dict(size=22, color="black"),
            title=None,
        ),
        legend=dict(
            font=dict(size=20, color="black"),
            title=dict(font=dict(size=20))
        ),
        width=800,
        height=700,
    )

    # Atualizando as barras
    fig.update_traces(
        textposition="outside",
        textfont=dict(size=22, color="black"),
        hovertemplate="<b>%{label}</b><br>Quantidade: %{value}<extra></extra>",
        hoverlabel=dict(font_size=18),
        marker=dict(line=dict(width=0.1), cornerradius=10)
    )

    return fig

def grafico_rank_alunos(dados):
    fig = px.bar(
        dados,
        x="N° empréstimos",
        y="Nome",
        orientation="h",
        title="Ranking de alunos",
        height=800,
        text="N° empréstimos",
        color_discrete_sequence=["#8E58BF"],
    )

    fig.update_layout(
        title=dict(font=dict(size=26)),
        xaxis=dict(
            title=dict(font=dict(size=22, color="black")),
            tickfont=dict(size=22, color="black"),
            tickmode="array",
            tickvals=dados["Nome"],
        ),
        yaxis=dict(
            tickfont=dict(size=22, color="black"),
            title=None,
        ),
        legend=dict(
            font=dict(size=20, color="black"),
            title=dict(font=dict(size=20))
        ),
        width=800,
        height=700,
    )

    fig.update_traces(
        textposition="outside",
        textfont=dict(size=22, color="black"),
        hovertemplate="<b>%{label}</b><br>Quantidade: %{value}<extra></extra>",
        hoverlabel=dict(font_size=18),
        marker=dict(line=dict(width=0.1), cornerradius=10)
    )

    return fig

def quebrar_linha(titulo, n_palavras=4):
    palavras = titulo.split()
    if len(palavras) > n_palavras:
        return ' '.join(palavras[:n_palavras]) + '<br>' + ' '.join(palavras[n_palavras:])
    return titulo

config = {
    "displayModeBar": "hover",
    "modeBarButtonsToRemove": [
        "zoom",
        "pan",
        "resetScale",
        "hover",
        "zoomIn",
        "zoomOut",
        "resetAxes",
        "lasso2d",
        "select2d",
        "zoom2d",
        "pan2d",
        "orbitRotation",
        "tableRotation",
        "save",
        "zoom2d",
        "sendDataToCloud",
        "toggleSpikelines",
        "toggleHover",
        "resetView",
        "toggleFullscreen",
    ],
    "displaylogo": False,
    "modeBarButtonsToAdd": ["downloadImage"],
}

custom_colors = [
    "#BEA6FF",
    "#AF54E4",
    "#A18AC2",
    "#8E74D2",
    "#751AAA",
    "#DCCEFF",
    "#A887FF",
    "#4E358E",
    "#7B7097",
    "#824BA7",
]

# Transformação de query com o id_escola sendo passado
status = execute_query(status, id_escola)
emprestimo_aluno = execute_query(emprestimo_aluno, id_escola)
datas_emp = execute_query(datas_emp, id_escola)
genero_m_emprestados = execute_query(genero_m_emprestados, id_escola)
qtd_livros = execute_query(qtd_livros, id_escola)
qtd_livro_emprestado = execute_query(qtd_livro_emprestado, id_escola)
livros_m_emprestados = execute_query(livros_m_emprestados, id_escola)
ranking_alunos = execute_query(ranking_alunos, id_escola)
livros_mais_emprestados = execute_query(livros_mais_emprestados_query, id_escola)
qtd_emprestimo_mes = execute_query(qtd_emprestimo_mes, id_escola)
livros_disponiveis = execute_query(livros_disponiveis, id_escola)
livros_geral = execute_query(livros_geral, id_escola)
quantidade_ano = execute_query(quantidade_ano, id_escola)
status_alunos = execute_query(status_alunos, id_escola)
acervo_geral = execute_query(acervo_geral, id_escola)
genero_total = execute_query(genero_total, id_escola)
livros_total = execute_query(livros_total, id_escola)
total_alunos = execute_query(total_alunos, id_escola)
pessoas_com_livro = execute_query(pessoas_com_livro, id_escola)
pessoas_cadastradas = execute_query(pessoas_cadastradas, id_escola)

# Rotulação dos meses
qtd_emprestimo_mes["Ano"] = qtd_emprestimo_mes["Ano"].astype(int)
qtd_emprestimo_mes["Mês"] = qtd_emprestimo_mes["Mês"].astype(int)
emprestimo_aluno["ano"] = emprestimo_aluno["ano"].astype(int)

qtd_emprestimo_mes["Data"] = pd.to_datetime(
    qtd_emprestimo_mes["Ano"].astype(str)
    + "-"
    + qtd_emprestimo_mes["Mês"].astype(str)
    + "-01",
    format="%Y-%m-%d",
)

emp_mes = qtd_emprestimo_mes.sort_values(by="Data", ascending=True)

# Mapeamento de meses
meses = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

meses_2 = {
    1: "Jan",
    2: "Fev",
    3: "Mar",
    4: "Abr",
    5: "Mai",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Set",
    10: "Out",
    11: "Nov",
    12: "Dez",
}

# Mapeando o número do mês para o nome do mês
emprestimo_aluno["mes"] = emprestimo_aluno["mes"].map(meses)
ranking_alunos["mes"] = ranking_alunos["mes"].map(meses)
livros_mais_emprestados["mes"] = livros_mais_emprestados["mes"].map(meses)
qtd_emprestimo_mes["Mês"] = qtd_emprestimo_mes["Mês"].map(meses)
emp_mes["Mês"] = emp_mes["Mês"].map(meses_2)
genero_m_emprestados["mes"] = genero_m_emprestados["mes"].map(meses)

# dicionario do status
dicionario = {"EM_DIA": "Em dia", "EM_ATRASO": "Em atraso"} # Colocar ATRASADO no lugar de EM_ATRASO

# renomeando os status do empréstimo
status = status.replace(dicionario)

# Extraindo os valores da consulta
total_livros_emprestados = livros_geral["total_livros_emprestados"].iloc[0]
livros_disponiveis = livros_geral["livros_disponiveis"].iloc[0]
total_geral = livros_geral["total_geral"].iloc[0]

# Cálculo do total de cada status
em_dia = status[status["Status"] == "Em dia"]["Quantidade"].sum()
em_atraso = status[status["Status"] == "Em atraso"]["Quantidade"].sum()

# Cálculo do acervo de livros disponiveis e emprestados
acervo_total = qtd_livros["quantidade"].sum()
emprestimos = qtd_livro_emprestado["quantidade"].sum()

# Converter a imagem para base64
image_path = "icone_download.svg"

with open(image_path, "rb") as image_file:
    encoded_image = base64.b64encode(image_file.read()).decode("utf-8")

# totalizador de pessoas cadastradas
tabela_cadastro = pessoas_cadastradas
totalizador_cadastro = pessoas_cadastradas["id"].count()
pessoas_cadastradas = pessoas_cadastradas[["Nome", "Telefone"]]

# totalizador de pessoas com livro
totalizador_pessoas_com_livro = pessoas_com_livro["id"].count()
tabela_pessoas_com_livro = pessoas_com_livro[["Nome", "Telefone"]]

# imagem de download
image_path2 = "icone_download_preto.svg"

with open(image_path2, "rb") as image_file:
    encoded_image2 = base64.b64encode(image_file.read()).decode("utf-8")

col1, col2, col3, col4 = st.columns([0.85, 2.4, 4, 1])

with col2:
    totalizador("Acervo geral", acervo_total, acervo_geral, "acervo_geral.xlsx")
    st.write("")
    st.write("")
    totalizador(
        "Pessoas cadastradas",
        totalizador_cadastro,
        pessoas_cadastradas,
        "pessoas_cadastradas.xlsx",
    )
    st.write("")
    st.write("")
    totalizador(
        "Pessoas com livros emprestados",
        totalizador_pessoas_com_livro,
        tabela_pessoas_com_livro,
        "pessoas_com_livro.xlsx",
    )

with col3:
    with st.container(height=501, border=True):

        st.write("")

        fig_status = px.pie(
            status,
            names="Status",
            values="Quantidade",
            hole=0.53,
            title="Livros emprestados",
            color="Status",
            color_discrete_map={"Em dia": "#D9D9D9", "Em atraso": "#CA3336"},
            height=400,
        )

        # Atualizando layout e legendas
        fig_status.update_layout(
            title_font_size=25,
            title_font_color="black",
            legend_font_size=20,
            legend_font_color="black",
            margin=dict(l=30, r=30, t=50, b=0),
            annotations=[
                {
                    "font": {"size": 65, "color": "black"},
                    "showarrow": False,
                    "text": f"{em_dia + em_atraso}",
                    "x": 0.5,
                    "y": 0.5,
                }
            ],
        )

        # Ajustando rótulos e hover
        fig_status.update_traces(
            textposition="auto",
            textfont_size=30,
            textfont_color="black",
            textinfo="value",
            hovertemplate="<b>%{percent}</b><br>%{label}<extra></extra>",
            hoverlabel_font_size=18,
        )

        st.plotly_chart(fig_status, use_container_width=True)

        tabela_download(status_alunos, "emprestimos.xlsx")

st.write("")
st.write("")

# #############

# Definir as opções de filtros (ano e mês)
ano_opcoes = emprestimo_aluno["ano"].unique()
mes_opcoes = emprestimo_aluno["mes"].unique()

st.divider()

# Lista com os meses
mes = mes_opcoes.tolist()

# Layout de filtros
col1, col2, col3, col4, col5, col6 = st.columns(6)

with col3:
    # Filtro de ano
    st.markdown(
        "<h2 style='font-size: 20px; font-weight: bold; margin-top: -18px; position: relative; padding-bottom: 5px;'>Selecione o ano:</h2>",
        unsafe_allow_html=True,
    )
    ano_selecionado = st.selectbox(
        "Selecione o ano", ano_opcoes, label_visibility="collapsed"
    )
    st.write("")

with col4:
    st.markdown(
        "<h2 style='font-size: 20px; font-weight: bold; margin-top: -18px; position: relative; padding-bottom: 5px;'>Selecione o mês:</h2>",
        unsafe_allow_html=True,
    )

    meses_disponiveis = (
        emprestimo_aluno[emprestimo_aluno["ano"] == ano_selecionado]["mes"]
        .unique()
        .tolist()
    )

    mes_selecionado = st.multiselect(
        "Selecione o mês",
        meses_disponiveis,
        label_visibility="collapsed",
        placeholder="Todos",
    )
    st.write("")

# Filtragem do DataFrame com base nas seleções
if ano_selecionado and mes_selecionado:
    if "Todos os meses" in mes_selecionado:
        filtro = emprestimo_aluno[emprestimo_aluno["ano"] == ano_selecionado]
    else:
        filtro = emprestimo_aluno[
            (emprestimo_aluno["ano"] == ano_selecionado)
            & (emprestimo_aluno["mes"].isin(mes_selecionado))
        ]

nuvem1, nuvem2, nuvem3 = st.columns([1, 6, 1])
with nuvem2:
    wordcloud, livros_mais_emprestados_mes = gerar_nuvem_emprestimos(
        emprestimo_aluno, ano_selecionado, mes_selecionado
    )

    if not mes_selecionado:
        st.subheader(f"Livros mais emprestados em {ano_selecionado}")
    else:
        st.subheader(f"Livros mais emprestados no período selecionado")

    fig, ax = plt.subplots(figsize=(10, 7))
    ax.imshow(wordcloud, interpolation="bilinear")
    ax.set_axis_off()
    st.pyplot(fig)
    st.write("")

# Emprestimos por mês/ano pelo ano selecionado
emprestimo_ano = emp_mes[emp_mes["Ano"] == ano_selecionado]

fig1 = plot_emprestimos_mensais()

with st.container(border=True):
    st.plotly_chart(fig1, use_container_width=True, config=config)


# Verificando se mes_selecionado está vazio ou se "Todos os meses" foi selecionado
if mes_selecionado and mes_selecionado != "Todos os meses":
    generos_filtrados = genero_m_emprestados[
        (genero_m_emprestados["ano"] == ano_selecionado)
        & (
            genero_m_emprestados["mes"].isin(mes_selecionado)
        )
    ]

    genero_total_filtrado = (
        generos_filtrados.groupby("Gênero")["Quantidade"].sum().reset_index()
    )
    genero_total_filtrado = genero_total_filtrado.sort_values(
        by="Quantidade", ascending=False
    )

    fig2 = grafico_emprestimo_genero(custom_colors)
    
else:
    generos_filtrados = genero_m_emprestados[
        genero_m_emprestados["ano"] == ano_selecionado
    ]

    genero_total_filtrado = (
        generos_filtrados.groupby("Gênero")["Quantidade"].sum().reset_index()
    )
    
    genero_total_filtrado = genero_total_filtrado.sort_values(
        by="Quantidade", ascending=False
    )

    fig2 = grafico_emprestimo_genero(custom_colors)

with st.container(border=True):
    st.plotly_chart(fig2, use_container_width=True, config=config)


# Total de livros pelo ano selecionado
livros_total_filtrado_ano = livros_total[livros_total["ano"] == ano_selecionado]
livros_total_filtrado_ano["mes"] = livros_total_filtrado_ano["mes"].map(meses)

coluna1, coluna2 = st.columns(2)
with coluna1:
    if "Todos os meses" in mes_selecionado or not mes_selecionado:
        livros_total_filtrado = livros_total_filtrado_ano.rename(
            columns={"titulo": "Título", "quantidade_emprestimos": "N° empréstimos"}
        )
        
        livros_total_filtrado = (
            livros_total_filtrado.groupby("Título")["N° empréstimos"]
            .sum()
            .reset_index()
        )
        
        livros_tabela = livros_total_filtrado
        
        livros_total_filtrado = livros_total_filtrado[["Título", "N° empréstimos"]].nlargest(10, "N° empréstimos")
        
        livros_total_filtrado = livros_total_filtrado.sort_values("N° empréstimos")
        
        livros_total_filtrado['Título'] = livros_total_filtrado['Título'].apply(quebrar_linha)
        
        fig3 = grafico_rank_livros(livros_total_filtrado)
        
        with st.container(height=780, border=True):
            st.plotly_chart(fig3, use_container_width=True, config=config)
            tabela_download(livros_tabela, "ranking_livros.xlsx")
            
    else:
        livros_total_filtrado_mes_selecionado = livros_total_filtrado_ano[
            livros_total_filtrado_ano["mes"].isin(mes_selecionado)
        ]
        
        livros_total_filtrado_mes_selecionado = (
            livros_total_filtrado_mes_selecionado.rename(
                columns={"titulo": "Título", "quantidade_emprestimos": "N° empréstimos"}
            )
        )

        livros_total_filtrado_mes_selecionado = (
            livros_total_filtrado_mes_selecionado.groupby("Título")["N° empréstimos"]
            .sum()
            .reset_index()
        )
        
        tabela_livros = livros_total_filtrado_mes_selecionado
        
        livros_total_filtrado_mes_selecionado = (
            livros_total_filtrado_mes_selecionado.nlargest(10, "N° empréstimos")
        )
        
        livros_total_filtrado_mes_selecionado = (
            livros_total_filtrado_mes_selecionado.sort_values("N° empréstimos")
        )
        
        livros_total_filtrado_mes_selecionado['Título'] = livros_total_filtrado_mes_selecionado['Título'].apply(quebrar_linha)

        fig3 = grafico_rank_livros(livros_total_filtrado_mes_selecionado)
        
        with st.container(height=780, border=True):
            st.plotly_chart(fig3, use_container_width=True, config=config)
            tabela_download(tabela_livros, "ranking_livros.xlsx")

total_alunos_filtrado_ano_completo = total_alunos[total_alunos["ano"] == ano_selecionado]

total_alunos_filtrado_ano_completo["mes"] = total_alunos_filtrado_ano_completo["mes"].map(meses)

with coluna2:
    if "Todos os meses" in mes_selecionado or not mes_selecionado:
        total_alunos_filtrado = total_alunos_filtrado_ano_completo.rename(
            columns={"nome": "Nome", "quantidade_emprestimos": "N° empréstimos"}
        )

        total_alunos_filtrado = (
            total_alunos_filtrado.groupby("Nome")["N° empréstimos"].sum().reset_index()
        )
        
        tabela_alunos = total_alunos_filtrado
        
        total_alunos_filtrado = total_alunos_filtrado.nlargest(10, "N° empréstimos") 
        
        tabela_total_alunos = total_alunos_filtrado
        
        total_alunos_filtrado = total_alunos_filtrado.sort_values("N° empréstimos")
        
        fig4 = grafico_rank_alunos(total_alunos_filtrado)

        with st.container(height=780, border=True):
            st.plotly_chart(fig4, use_container_width=True, config=config)
            tabela_download(tabela_alunos, "ranking_alunos.xlsx")

    else:
        total_alunos_filtrado_mes_selecionado = total_alunos_filtrado_ano_completo[
            total_alunos_filtrado_ano_completo["mes"].isin(mes_selecionado)
        ]
        total_alunos_filtrado_mes_selecionado = (
            total_alunos_filtrado_mes_selecionado.rename(
                columns={"nome": "Nome", "quantidade_emprestimos": "N° empréstimos"}
            )
        )

        total_alunos_filtrado_mes_selecionado = (
            total_alunos_filtrado_mes_selecionado.groupby("Nome")["N° empréstimos"]
            .sum()
            .reset_index()
        )
        
        tabela_alunos = total_alunos_filtrado_mes_selecionado
        total_alunos_filtrado_mes_selecionado = (
            total_alunos_filtrado_mes_selecionado.nlargest(10, "N° empréstimos")
        )
        
        total_alunos_filtrado_mes_selecionado = (
            total_alunos_filtrado_mes_selecionado.sort_values("N° empréstimos")
        )
        
        fig4 = grafico_rank_alunos(total_alunos_filtrado_mes_selecionado)

        with st.container(height=780, border=True):
            st.plotly_chart(fig4, use_container_width=True, config=config)
            tabela_download(tabela_alunos, "ranking_alunos.xlsx")
